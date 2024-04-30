import cx_Oracle
import os
import pandas as pd
from openpyxl import Workbook, load_workbook


class Clase_PretupsVsSap:
    def __init__(self, fecha_inicio, fecha_fin, fecha_hoy, user, password, database, ruta_ccd):
        self.fecha_inicio = fecha_inicio
        self.fecha_fin = fecha_fin
        self.fecha_hoy = fecha_hoy
        self.user = user
        self.password = password
        self.database = database
        self.ruta_ccd = ruta_ccd
        self.paCargaRecargas = (
            """BEGIN CCD_CARGA_DATA_RECARGA('"""+self.fecha_inicio+"""','"""+self.fecha_fin+"""');
            END;"""
            )
        self.paCargaConciliacion = (
            """BEGIN C26823.CCD_ACTUALIZA_CONCILIACION;
            END;"""
            )
        self.ruta_file_IPE_PRETUPS = self.ruta_ccd + "/REPORTES MENSUALES/REPORTES DE CONTROL/Reportes SOX/C19-Pretups vs SAP_Henry y Diego/IPE/2.-ORACLE (PRETUPS)"
        self.ruta_file_reporte_final = self.ruta_ccd + "/REPORTES MENSUALES/REPORTES DE CONTROL/Reportes SOX/C19-Pretups vs SAP_Henry y Diego/REPORTE FINAL"
        self.excel_file_BASES = self.ruta_ccd + "/REPORTES MENSUALES/REPORTES DE CONTROL/Reportes SOX/C19-Pretups vs SAP_Henry y Diego/BASES.xlsx"
        self.excel_file_BASES_MARGEN = self.ruta_ccd + "/REPORTES MENSUALES/REPORTES DE CONTROL/Reportes SOX/C19-Pretups vs SAP_Henry y Diego/MARGEN.xlsx"
        self.hojas_y_tablas = [("SAP", "TEMP_FACTURAS_SAP"), ("MARGEN", "TEMP_MARGEN")]
        self.excel_file_Conciliacion = self.ruta_ccd + "/REPORTES MENSUALES/REPORTES DE CONTROL/Reportes SOX/C19-Pretups vs SAP_Henry y Diego/Conciliación PRETUPS-SAP.xlsx"
        self.getRecargasPretups = """SELECT * FROM C26823.TEMP_PRETUPS"""
        self.sql_update_temp_facturas_sap = ("""UPDATE TEMP_FACTURAS_SAP SET PERIODO = '"""+self.fecha_hoy+"""'""")
        self.sql_insert_tabla_consolidada = (
            """INSERT INTO USRCCD.CONCILIACION_SAP_PRETUPS
            SELECT FECHA, DEUDOR, NOMBRE, REGION_, TIPO, ANALISTA, FACTURA, IMPORTE_SAP, IMPORTE_PRETUPS, COMISION,  IMPORTE_FINAL, DIFERENCIA, COMENTARIOS, PERIODO  FROM TEMP_FACTURAS_SAP"""
            )
        self.sql_delete_consolidado = (
            """DELETE FROM USRCCD.CONCILIACION_SAP_PRETUPS
            WHERE PERIODO = '"""+self.fecha_hoy+"""'"""
        )
        self.sql_select_reporte_final = (
            """SELECT * FROM USRCCD.CONCILIACION_SAP_PRETUPS
            WHERE PERIODO ='"""+self.fecha_hoy+"""'"""
        )

    def format_date(self, date_str):
        # Convierte la fecha en una cadena con formato DD/MM/YYYY
        day, month, year = date_str.split(".")
        return f"{day}/{month}/{year}"

    def extract_data_xlsx(self, file_path):
        df = pd.read_excel(file_path)
        # Encontrar los índices de las primeras ocurrencias de las palabras clave
        first_margen_idx = df[df["Unnamed: 1"] == "Margen % (ZD03)"].index[0]
        first_cliente_idx = df[df["Clave del registro de condición"] == "Cliente"].index[0]
        first_fecha_idx = df[df["Clave del registro de condición"] == "Período de validez"].index[0]
        
        # Número de filas por registro
        rows_per_record = 14
        # Extraer los datos y almacenarlos en listas
        clientes = [df.iloc[i, 2] for i in range(first_cliente_idx, df.shape[0], rows_per_record + 1)]
        fechas_i = [self.format_date(df.iloc[i, 2]) for i in range(first_fecha_idx, df.shape[0], rows_per_record + 1)]
        fechas_f = [self.format_date(df.iloc[i, 5]) for i in range(first_fecha_idx, df.shape[0], rows_per_record + 1)]
        margenes = [df.iloc[i, 4] for i in range(first_margen_idx, df.shape[0], rows_per_record + 1)]
        # Crear y retornar un DataFrame con los datos extraídos
        margen_pd = pd.DataFrame({"PADRE": clientes, "FECHA_I": fechas_i, "FECHA_F": fechas_f, "MARGEN": margenes,})
        margen_pd["COMISION"] = ""
        return margen_pd

    def cargaBasesF(self):
        connection = self.connectionOPERCOM()
        cursor = connection.cursor()
        workbook = load_workbook(self.excel_file_BASES)

        for sheet_name, table_name in self.hojas_y_tablas:
            sheet = workbook[sheet_name]
            # Obtén los datos de la hoja Excel y conviértelos a una lista de listas
            excel_data = [list(row) for row in sheet.iter_rows(min_row=2, values_only=True)]
            # Define la consulta INSERT con la cantidad adecuada de marcadores de posición
            insert_query = f"INSERT INTO {table_name} VALUES ({','.join([':{}'.format(i+1) for i in range(len(excel_data[0]))])})"
            cursor.executemany(insert_query, excel_data)
            # print(excel_data)
            connection.commit()

        cursor.close()
        connection.close()

    def cargaBases(self):
        connection = self.connectionOPERCOM()
        cursor = connection.cursor()

        xls = pd.ExcelFile(self.excel_file_BASES)
        df_SAP = xls.parse("SAP")
        column_SAP = ["Fecha factura", "Solicitante", "Nombre 1", "Factura", "Valor neto", "Impuesto", "Denominación", "Tipo de documento co", "Factura esta anulada",]
        df_MARGEN = self.extract_data_xlsx(self.excel_file_BASES_MARGEN)
        column_margen = ["PADRE", "FECHA_I", "FECHA_F", "COMISION", "MARGEN"]
        df_SAP = df_SAP[column_SAP]
        df_MARGEN = df_MARGEN[column_margen]
        df_MARGEN["MARGEN"] = df_MARGEN["MARGEN"].astype(float, 4)
        df_MARGEN["COMISION"] = df_MARGEN["MARGEN"]

        # LIMPIAR DATA
        valores_permitidos = ["SERVICIO RECARGA VIRTUAL STREET SELLER"]
        df_SAP = df_SAP[df_SAP["Denominación"].isin(valores_permitidos)]
        valores_permitidos = ["Factura"]
        df_SAP = df_SAP[df_SAP["Tipo de documento co"].isin(valores_permitidos)]
        valores_permitidos = ["No"]
        df_SAP = df_SAP[df_SAP["Factura esta anulada"].isin(valores_permitidos)]
        column_SAP = ["Fecha factura", "Solicitante", "Nombre 1", "Factura", "Valor neto", "Impuesto",]
        df_SAP = df_SAP[column_SAP]

        for i in range(10):
            df_SAP.insert(1, "columnName" + str(i + 1), "")

        column_SAP = ["Fecha factura", "Solicitante", "Nombre 1", "columnName1", "columnName2", "columnName3", "Factura", "Valor neto", "Impuesto", "columnName4", "columnName5", "columnName6", "columnName7", "columnName8", "columnName9", "columnName10",]
        df_SAP = df_SAP[column_SAP]
        rows = [tuple(x) for x in df_SAP.to_numpy()]
        rows_with_date_SAP = [row for row in rows]
        rows = [tuple(x) for x in df_MARGEN.to_numpy()]
        rows_with_date_MARGEN = [row for row in rows]

        insert_sql_SAP = f"INSERT INTO TEMP_FACTURAS_SAP VALUES ({','.join([':{}'.format(i+1) for i in range(len(rows_with_date_SAP[0]))])})"
        cursor.executemany(insert_sql_SAP, rows_with_date_SAP)
        # print(rows_with_date_MARGEN)
        insert_sql_MARGEN = f"INSERT INTO TEMP_MARGEN VALUES ({','.join([':{}'.format(i+1) for i in range(len(rows_with_date_MARGEN[0]))])})"
        cursor.executemany(insert_sql_MARGEN, rows_with_date_MARGEN)

        connection.commit()
        cursor.close()
        connection.close()

    def conciliacion(self):
        connection = self.connectionOPERCOM()
        cursor = connection.cursor()
        cursor.execute(self.paCargaConciliacion)
        cursor.execute(self.sql_update_temp_facturas_sap)
        cursor.execute(self.sql_insert_tabla_consolidada)
        connection.commit()
        cursor.close()
        connection.close()

    def cargaPretupsRecargas(self):
        connection = self.connectionOPERCOM()
        cursor = connection.cursor()
        cursor.execute(self.sql_delete_consolidado)
        cursor.execute(self.paCargaRecargas)
        connection.commit()
        cursor.close()
        connection.close()

    def connectionOPERCOM(self):
        dsn = cx_Oracle.makedsn("scan-fcprod.tim.com.pe", "1521", service_name=self.database)
        connection = cx_Oracle.connect(user=self.user, password=self.password, dsn=dsn)
        return connection

    def exportHojasPretupsSap(self, fecha):
        connection = self.connectionOPERCOM()
        cursor = connection.cursor()
        cursor.execute(self.sql_select_reporte_final)
        excel_file_IPE = self.ruta_file_reporte_final+"/Conciliación PRETUPS-SAP_"+fecha+".xlsx"

        column_names = [col[0] for col in cursor.description]
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(column_names)
        for row in cursor:
            sheet.append(row)

        workbook.save(excel_file_IPE)
        cursor.close()
        connection.close()

    def exportIPEpretups(self, fecha):
        connection = self.connectionOPERCOM()
        cursor = connection.cursor()
        cursor.execute(self.getRecargasPretups)
        excel_file_IPE = self.ruta_file_IPE_PRETUPS+"/PRETUPS "+fecha+".xlsx"

        column_names = [col[0] for col in cursor.description]
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(column_names)
        for row in cursor:
            sheet.append(row)

        workbook.save(excel_file_IPE)
        cursor.close()
        connection.close()
