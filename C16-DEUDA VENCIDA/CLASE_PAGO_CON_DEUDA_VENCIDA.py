import cx_Oracle
import pandas as pd
from openpyxl import Workbook, load_workbook


class Pago_con_deuda_vencida:
    def __init__(self, fecha_hoy, fecha_hoy_date, user, password, database, ruta_ccd,):
        self.fecha_hoy = fecha_hoy
        self.fecha_hoy_date = fecha_hoy_date
        self.condicion = "SI"
        self.user = user
        self.password = password
        self.database = database
        self.ruta_ccd = ruta_ccd

        self.paActualizaPagoDeComisiones = (
            """BEGIN C26823.CCD_ACTUALIZA_PAGO_COMISIONES('""" + self.fecha_hoy + """', '""" + self.fecha_hoy_date + """');
            END;"""
            )

        self.paALimpiaTablas = (
            """BEGIN C26823.CCD_LIMPIAR_TABLAS('""" + fecha_hoy + """');
            END;"""
            )

        self.sql_insert_tabla_historica = """INSERT INTO USRCCD.PAGO_DEUDA_VENCIDA_FINAL SELECT * FROM DIN_PAGO_DEUDA_VENCIDA_FINAL"""

        self.sql_select_reporte_final = (
            """SELECT * FROM USRCCD.PAGO_DEUDA_VENCIDA_FINAL 
            WHERE FLAG = '""" + self.condicion + """' AND PERIODO ='""" + self.fecha_hoy + """'"""
            )

        self.hojas_y_tablas = [
            ("FBL5N", "C26823.TEMP_ABONOS"),
            ("FBL1N BLOQUEADAS", "C26823.TEMP_FBL1N"),
            ("FBL1N PAGADAS", "C26823.TEMP_PAGO_DEUDA_VENCIDA_FINAL"),
            ("ZFIR60", "C26823.TEMP_ZFIR"),
        ]

        self.ruta_file_reporte_final = (self.ruta_ccd + "REPORTES MENSUALES\REPORTES DE CONTROL/Reportes SOX/C15-Pago con deuda vencida_Henry/REPORTE FINAL")
        self.excel_file_BASES = (self.ruta_ccd + "REPORTES MENSUALES/REPORTES DE CONTROL/Reportes SOX/C15-Pago con deuda vencida_Henry/BASES.xlsx")


    def cargaBases(self):
        connection = self.connectionOPERCOM()
        cursor = connection.cursor()

        xls = pd.ExcelFile(self.excel_file_BASES)

        df_FBL5N = xls.parse("FBL5N")
        column_FBL5N = ["Cuenta", "Clase", "Importe en ML"]

        df_FBL1N_BLOQUEADAS = xls.parse("FBL1N BLOQUEADAS")
        column_FBL1N_BLOQUEADAS = ["Cuenta", "Fecha doc.", "BP", "Importe en ML"]

        df_ZFIR60 = xls.parse("ZFIR60")
        column_ZFIR60 = ["Cliente Pa", "Área Ctrl.", "Total Venc"]

        df_FBL1N_PAGADAS = xls.parse("FBL1N PAGADAS")
        column_FBL1N_PAGADAS = ["Fecha pago", "Cuenta", "Clase", "Importe en ML"]

        df_FBL5N = df_FBL5N[column_FBL5N]
        df_FBL1N_BLOQUEADAS = df_FBL1N_BLOQUEADAS[column_FBL1N_BLOQUEADAS]
        df_FBL1N_PAGADAS = df_FBL1N_PAGADAS[column_FBL1N_PAGADAS]
        df_ZFIR60 = df_ZFIR60[column_ZFIR60]

        # LIMPIAR DATA
        df_FBL1N_BLOQUEADAS = df_FBL1N_BLOQUEADAS.dropna(subset=["BP"])

        valores_permitidos = ["J9"]
        df_FBL5N = df_FBL5N[df_FBL5N["Clase"].isin(valores_permitidos)]
        
        # FILTRAR NEGATIVOS
        valores_permitidos = ["PE01", "PE02", "PE03", "PE04", "PE05", "PE07", "PE09", "PE10", "PE11", "PE12", "PE13", "PE14", "PE15", "PE16", "PE17",]
        df_ZFIR60 = df_ZFIR60[df_ZFIR60["Área Ctrl."].isin(valores_permitidos)]

        for i in range(3):
            df_FBL5N.insert(1, "columnName" + str(i + 1), "")
            df_FBL1N_BLOQUEADAS.insert(1, "columnName" + str(i + 1), "")

        df_ZFIR60.insert(1, "columnName1", "")

        column_FBL5N = ["columnName1", "Cuenta", "Clase", "columnName2", "Importe en ML", "columnName3",]
        df_FBL5N = df_FBL5N[column_FBL5N]
        
        column_FBL1N_BLOQUEADAS = ["columnName1", "Cuenta", "columnName2", "Fecha doc.", "BP", "Importe en ML", "columnName3",]
        df_FBL1N_BLOQUEADAS = df_FBL1N_BLOQUEADAS[column_FBL1N_BLOQUEADAS]

        column_ZFIR60 = ["columnName1", "Cliente Pa", "Área Ctrl.", "Total Venc"]
        df_ZFIR60 = df_ZFIR60[column_ZFIR60]

        rows = [tuple(x) for x in df_FBL5N.to_numpy()]
        rows_with_date_FBL5N = [row for row in rows]
        rows = [tuple(x) for x in df_FBL1N_BLOQUEADAS.to_numpy()]
        rows_with_date_FBL1N = [row for row in rows]
        rows = [tuple(x) for x in df_ZFIR60.to_numpy()]
        rows_with_date_ZFIR60 = [row for row in rows]
        rows = [tuple(x) for x in df_FBL1N_PAGADAS.to_numpy()]
        rows_with_date_FBL1N_PAGADAS = [row for row in rows]

        insert_sql_FBL5N = f"INSERT INTO C26823.TEMP_ABONOS VALUES ({','.join([':{}'.format(i+1) for i in range(len(rows_with_date_FBL5N[0]))])})"
        insert_sql_FBL1N_BLOQUEADAS = f"INSERT INTO C26823.TEMP_FBL1N VALUES ({','.join([':{}'.format(i+1) for i in range(len(rows_with_date_FBL1N[0]))])})"
        insert_sql_FBL1N_PAGADAS = f"INSERT INTO C26823.TEMP_PAGO_DEUDA_VENCIDA_FINAL VALUES ({','.join([':{}'.format(i+1) for i in range(len(rows_with_date_FBL1N_PAGADAS[0]))])})"
        insert_sql_ZFIR60 = f"INSERT INTO C26823.TEMP_ZFIR VALUES ({','.join([':{}'.format(i+1) for i in range(len(rows_with_date_ZFIR60[0]))])})"

        cursor.executemany(insert_sql_FBL5N, rows_with_date_FBL5N)
        cursor.executemany(insert_sql_FBL1N_BLOQUEADAS, rows_with_date_FBL1N)
        cursor.executemany(insert_sql_FBL1N_PAGADAS, rows_with_date_FBL1N_PAGADAS)
        cursor.executemany(insert_sql_ZFIR60, rows_with_date_ZFIR60)

        connection.commit()
        cursor.close()
        connection.close()

    def cargaBasesf(self):
        connection = self.connectionOPERCOM()
        cursor = connection.cursor()
        workbook = load_workbook(self.excel_file_BASES)

        for sheet_name, table_name in self.hojas_y_tablas:
            sheet = workbook[sheet_name]

            # Obtén los datos de la hoja Excel y conviértelos a una lista de listas
            excel_data = [list(row) for row in sheet.iter_rows(min_row = 2, values_only = True)]

            # Define la consulta INSERT con la cantidad adecuada de marcadores de posición
            insert_query = f"INSERT INTO {table_name} VALUES ({','.join([':{}'.format(i+1) for i in range(len(excel_data[0]))])})"
            cursor.executemany(insert_query, excel_data)
            connection.commit()
            
        cursor.close()
        connection.close()

    def ActualizaPagoDeComisiones(self):
        connection = self.connectionOPERCOM()
        cursor = connection.cursor()
        cursor.execute(self.paActualizaPagoDeComisiones)
        cursor.execute(self.sql_insert_tabla_historica)
        connection.commit()
        cursor.close()
        connection.close()

    def limpiarTablas(self):
        connection = self.connectionOPERCOM()
        cursor = connection.cursor()
        cursor.execute(self.paALimpiaTablas)
        connection.commit()
        cursor.close()
        connection.close()

    def connectionOPERCOM(self):
        dsn = cx_Oracle.makedsn("scan-fcprod.tim.com.pe", "1521", service_name = self.database)
        connection = cx_Oracle.connect(user = self.user, password = self.password, dsn = dsn)
        return connection

    def exportHojas(self, fecha):
        connection = self.connectionOPERCOM()
        cursor = connection.cursor()
        cursor.execute(self.sql_select_reporte_final)
        excel_file_IPE = (self.ruta_file_reporte_final + "/Pago_Deuda_Vencida_" + fecha + ".xlsx")
        column_names = [col[0] for col in cursor.description]
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(column_names)
        for row in cursor:
            sheet.append(row)
        workbook.save(excel_file_IPE)
        cursor.close()
        connection.close()
